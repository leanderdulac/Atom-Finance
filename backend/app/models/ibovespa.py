"""
Ibovespa 18-Asset Universe + RL Portfolio Optimizer
=====================================================
Algoritmo de otimização: Cross-Entropy Method (CEM) — método de busca de
política estocástica amplamente usado em RL (AlphaZero, OpenAI ES, etc.).

Pipeline:
1. GBM correlacionado gera N trajetórias para os 18 ativos.
2. CEM amostra vetores de pesos W ~ N(μ, Σ), avalia cada W por função
   de recompensa (Sharpe penalizado pelo perfil) e atualiza (μ, Σ) com
   os W elite.  Repete K iterações até convergência.
3. Retorna pesos ótimos + métricas do portfólio otimizado.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from functools import partial
from typing import Literal

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.data_fetcher import DataFetcher

logger = logging.getLogger(__name__)

# Last successful refresh timestamp
_LAST_REFRESH: datetime | None = None
_REFRESH_INTERVAL = timedelta(hours=1)


# ── 18 Ibovespa assets ────────────────────────────────────────────────────────

IBOVESPA_ASSETS: list[dict] = [
    {"ticker": "PETR4", "S0": 40.0,  "mu": 0.12, "sigma": 0.38, "sector": "Energia"},
    {"ticker": "VALE3", "S0": 65.0,  "mu": 0.10, "sigma": 0.32, "sector": "Mineração"},
    {"ticker": "ITUB4", "S0": 35.0,  "mu": 0.10, "sigma": 0.26, "sector": "Financeiro"},
    {"ticker": "BBDC4", "S0": 15.0,  "mu": 0.08, "sigma": 0.28, "sector": "Financeiro"},
    {"ticker": "ABEV3", "S0": 14.0,  "mu": 0.07, "sigma": 0.20, "sector": "Consumo"},
    {"ticker": "WEGE3", "S0": 50.0,  "mu": 0.15, "sigma": 0.30, "sector": "Industrial"},
    {"ticker": "RENT3", "S0": 55.0,  "mu": 0.12, "sigma": 0.34, "sector": "Mobilidade"},
    {"ticker": "BBAS3", "S0": 28.0,  "mu": 0.09, "sigma": 0.25, "sector": "Financeiro"},
    {"ticker": "RDOR3", "S0": 30.0,  "mu": 0.13, "sigma": 0.32, "sector": "Saúde"},
    {"ticker": "HAPV3", "S0":  6.0,  "mu": 0.08, "sigma": 0.44, "sector": "Saúde"},
    {"ticker": "PRIO3", "S0": 42.0,  "mu": 0.16, "sigma": 0.42, "sector": "Energia"},
    {"ticker": "GGBR4", "S0": 20.0,  "mu": 0.09, "sigma": 0.31, "sector": "Siderurgia"},
    {"ticker": "CSAN3", "S0": 16.0,  "mu": 0.10, "sigma": 0.35, "sector": "Energia"},
    {"ticker": "JBSS3", "S0": 22.0,  "mu": 0.11, "sigma": 0.28, "sector": "Alimentos"},
    {"ticker": "MGLU3", "S0":  4.0,  "mu": 0.14, "sigma": 0.58, "sector": "Varejo"},
    {"ticker": "LREN3", "S0": 18.0,  "mu": 0.09, "sigma": 0.30, "sector": "Varejo"},
    {"ticker": "SUZB3", "S0": 60.0,  "mu": 0.10, "sigma": 0.27, "sector": "Papel/Celulose"},
    {"ticker": "CSNA3", "S0": 14.0,  "mu": 0.08, "sigma": 0.34, "sector": "Siderurgia"},
]

# Ibovespa approximate correlation (sector-blocked structure)
# Financial: ITUB4, BBDC4, BBAS3 → high intra-sector
# Energy:    PETR4, PRIO3, CSAN3
# Mining/Steel: VALE3, GGBR4, CSNA3
# Consumer/Health low correlation with commodities
def _build_correlation_matrix() -> np.ndarray:
    n = len(IBOVESPA_ASSETS)
    tickers = [a["ticker"] for a in IBOVESPA_ASSETS]
    C = np.eye(n)
    # sector groups → high intra-sector correlation
    sectors = {a["ticker"]: a["sector"] for a in IBOVESPA_ASSETS}
    for i in range(n):
        for j in range(i + 1, n):
            ti, tj = tickers[i], tickers[j]
            if sectors[ti] == sectors[tj]:
                rho = 0.65       # same sector
            elif any(x in sectors[ti] for x in ["Energia", "Mineração", "Siderurgia"]) and \
                 any(x in sectors[tj] for x in ["Energia", "Mineração", "Siderurgia"]):
                rho = 0.40       # commodity cluster
            elif any(x in sectors[ti] for x in ["Financeiro"]) and \
                 any(x in sectors[tj] for x in ["Financeiro"]):
                rho = 0.70
            else:
                rho = 0.25       # diversified
            C[i, j] = C[j, i] = rho
    # ensure positive semi-definite by Higham nearest PD
    eigvals, eigvecs = np.linalg.eigh(C)
    eigvals = np.clip(eigvals, 1e-6, None)
    C = eigvecs @ np.diag(eigvals) @ eigvecs.T
    # re-normalise diagonal to 1
    d = np.sqrt(np.diag(C))
    C = C / np.outer(d, d)
    return C

CORR_MATRIX = _build_correlation_matrix()


async def refresh_ibovespa_params(force: bool = False) -> bool:
    """
    Update IBOVESPA_ASSETS prices (S0) using real-time data from Yahoo Finance.
    """
    global _LAST_REFRESH
    now = datetime.now()
    
    if not force and _LAST_REFRESH and (now - _LAST_REFRESH) < _REFRESH_INTERVAL:
        return False

    tickers = [a["ticker"] for a in IBOVESPA_ASSETS]
    yf_tickers = [f"{t}.SA" for t in tickers]   # Yahoo Finance B3 suffix
    
    logger.info("Initializing Ibovespa parameters sync with Yahoo Finance...")
    
    try:
        quotes = await DataFetcher.get_multiple_quotes_async(yf_tickers)
        if not quotes:
            logger.warning("Dynamic refresh failed: No quotes returned from Yahoo Finance.")
            return False

        updated_count = 0
        from app.models.kronos_agent import KronosAgent
        import asyncio

        for asset in IBOVESPA_ASSETS:
            ticker = asset["ticker"]
            yf_ticker = f"{ticker}.SA"
            
            # 1. Update S0 (Price)
            if yf_ticker in quotes and quotes[yf_ticker]:
                price = quotes[yf_ticker].get("price")
                if price and price > 0:
                    asset["S0"] = price
            
            # 2. Update mu (Drift) via Kronos
            try:
                kronos_pred = await asyncio.to_thread(KronosAgent.predict, yf_ticker, 30)
                if kronos_pred:
                    # pred_ret is % return in 30 days (business days approx)
                    pred_ret = kronos_pred["predicted_return_pct"] / 100.0
                    ann_mu = pred_ret * (252 / 30) # Annualize roughly
                    # Limit extreme drift values to realistic bounds
                    ann_mu = max(-0.50, min(0.60, ann_mu))
                    asset["mu"] = round(ann_mu, 4)
                    updated_count += 1
            except Exception as e:
                logger.error("Failed to infer Kronos drift for %s: %s", ticker, e)

        _LAST_REFRESH = now
        logger.info("Successfully updated %d Ibovespa assets via real-time data & Kronos AI.", updated_count)
        return True
    except Exception as exc:
        logger.error("Error during Ibovespa parameter refresh: %s", exc)
        return False


# ── GBM multi-asset simulation ────────────────────────────────────────────────

def simulate_ibovespa(
    n_steps: int = 252,
    n_paths: int = 500,
    T: float = 1.0,
    seed: int | None = 42,
) -> dict:
    """
    Simulate correlated GBM for all 18 Ibovespa assets.
    Returns per-asset statistics (mean, p5, p95, mean_pct) + raw daily returns matrix.
    """
    if seed is not None:
        np.random.seed(seed)

    n = len(IBOVESPA_ASSETS)
    dt = T / n_steps
    L = np.linalg.cholesky(CORR_MATRIX)

    Z = np.random.standard_normal((n_paths, n_steps, n))
    W = Z @ L.T

    time_days = [int(round(t * 252)) for t in np.linspace(0, T, n_steps + 1)]
    asset_results = []
    # raw daily returns matrix (n_paths × n_steps × n_assets) for optimizer
    daily_returns = np.zeros((n_paths, n_steps, n))

    for i, asset in enumerate(IBOVESPA_ASSETS):
        S0, mu, sigma = asset["S0"], asset["mu"], asset["sigma"]
        inc = (mu - 0.5 * sigma ** 2) * dt + sigma * np.sqrt(dt) * W[:, :, i]
        daily_returns[:, :, i] = inc  # log-returns

        log_paths = np.concatenate([np.zeros((n_paths, 1)), np.cumsum(inc, axis=1)], axis=1)
        paths = S0 * np.exp(log_paths)

        mean_path = paths.mean(axis=0)
        p5  = np.percentile(paths, 5,  axis=0)
        p95 = np.percentile(paths, 95, axis=0)
        terminal = paths[:, -1]

        asset_results.append({
            "ticker":  asset["ticker"],
            "sector":  asset["sector"],
            "S0":      S0,
            "mu":      mu,
            "sigma":   sigma,
            "mean":    [round(float(v), 3) for v in mean_path],
            "p5":      [round(float(v), 3) for v in p5],
            "p95":     [round(float(v), 3) for v in p95],
            "mean_pct": [round(float((v / S0 - 1) * 100), 3) for v in mean_path],
            "terminal": {
                "mean":              round(float(terminal.mean()), 3),
                "std":               round(float(terminal.std()), 3),
                "p5":                round(float(np.percentile(terminal, 5)), 3),
                "p50":               round(float(np.median(terminal)), 3),
                "p95":               round(float(np.percentile(terminal, 95)), 3),
                "expected_return_pct": round(float((terminal.mean() / S0 - 1) * 100), 2),
                "prob_above_S0":     round(float((terminal > S0).mean()), 4),
            },
        })

    return {
        "T": T,
        "n_steps": n_steps,
        "n_paths": n_paths,
        "time_days": time_days,
        "assets": asset_results,
        "_daily_returns": daily_returns,   # internal — not serialised to JSON
    }


# ── Cross-Entropy Method (CEM) portfolio optimizer ────────────────────────────

@dataclass
class RLOptimizeResult:
    weights: np.ndarray
    tickers: list[str]
    iterations: int
    reward_history: list[float]
    portfolio_return: float
    portfolio_vol: float
    sharpe_ratio: float
    profile: str


def _reward(
    weights: np.ndarray,
    daily_returns: np.ndarray,  # (n_paths, n_steps, n_assets)  log-returns
    profile: Literal["conservador", "agressivo"],
) -> float:
    """
    Reward = Sharpe ratio with profile-specific penalty/bonus.

    Conservative: high Sharpe with vol penalty (prefer min-variance)
    Aggressive: high Sharpe with return bonus (prefer max-return)
    """
    # portfolio log-return per path per step
    port_log_returns = daily_returns @ weights          # (n_paths, n_steps)

    # annualised statistics (average across paths for mean/vol)
    daily_mean = port_log_returns.mean()
    daily_vol  = port_log_returns.std()

    ann_return = daily_mean * 252
    ann_vol    = daily_vol  * np.sqrt(252)
    sharpe     = ann_return / (ann_vol + 1e-8)

    if profile == "conservador":
        # penalise high volatility heavily
        return float(sharpe - 1.5 * ann_vol)
    else:
        # bonus for raw expected return
        return float(ann_return + 0.5 * sharpe)


def cem_optimize(
    daily_returns: np.ndarray,
    profile: Literal["conservador", "agressivo"],
    n_iterations: int = 60,
    n_samples: int = 300,
    elite_fraction: float = 0.20,
    max_weight: float = 0.25,   # max allocation per asset
    seed: int | None = 7,
) -> RLOptimizeResult:
    """
    Cross-Entropy Method for portfolio weight optimisation.

    Each iteration:
      1. Sample K weight vectors W_k ~ Dirichlet(alpha_k)
      2. Score each W_k with the reward function
      3. Keep the top elite_fraction
      4. Fit a new Dirichlet to the elite weights
    """
    if seed is not None:
        np.random.seed(seed)

    n = daily_returns.shape[2]
    n_elite = max(5, int(n_samples * elite_fraction))

    # Dirichlet concentration: start uniform
    alpha = np.ones(n)
    reward_history: list[float] = []
    best_weights = np.ones(n) / n

    for it in range(n_iterations):
        # 1. Sample
        W = np.random.dirichlet(alpha, size=n_samples)  # (n_samples, n)

        # 2. Enforce max_weight cap by renormalising
        W = np.clip(W, 0, max_weight)
        row_sums = W.sum(axis=1, keepdims=True)
        row_sums[row_sums == 0] = 1.0
        W = W / row_sums

        # 3. Score
        scores = np.array([_reward(W[k], daily_returns, profile) for k in range(n_samples)])

        # 4. Elite selection
        elite_idx = np.argsort(scores)[-n_elite:]
        elite_W = W[elite_idx]

        best_idx = int(np.argmax(scores))
        best_weights = W[best_idx]
        reward_history.append(float(scores[best_idx]))

        # 5. Update Dirichlet alpha (MLE for Dirichlet from elite samples)
        elite_mean = elite_W.mean(axis=0)
        elite_var  = elite_W.var(axis=0)
        # alpha_new = mean * (mean*(1-mean)/var - 1)   (method of moments)
        ratio = np.maximum(1e-6, elite_mean * (1 - elite_mean) / (elite_var + 1e-12) - 1)
        alpha = elite_mean * ratio
        alpha = np.clip(alpha, 0.1, 100)   # numerical stability

    # Final portfolio stats using best weights
    port_returns = (daily_returns @ best_weights).mean(axis=0)  # mean across paths
    ann_return = float(port_returns.mean() * 252)
    ann_vol    = float(port_returns.std() * np.sqrt(252))
    sharpe     = ann_return / (ann_vol + 1e-8)

    return RLOptimizeResult(
        weights=best_weights,
        tickers=[a["ticker"] for a in IBOVESPA_ASSETS],
        iterations=n_iterations,
        reward_history=reward_history,
        portfolio_return=ann_return,
        portfolio_vol=ann_vol,
        sharpe_ratio=sharpe,
        profile=profile,
    )


# ── Excel report generator ────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")
_HDR_FONT  = Font(color="FFFFFF", bold=True)
_GREEN_FG  = Font(color="22C55E", bold=True)
_RED_FG    = Font(color="EF4444", bold=True)
_GOLD_FILL = PatternFill("solid", fgColor="2D2D1A")


def _header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def generate_excel_report(
    sim: dict,
    rl_result: RLOptimizeResult,
    initial_capital: float,
    profile: str,
) -> bytes:
    """
    Generate a 3-sheet Excel workbook:
      1. Portfólio RL   — weights, allocation, expected return per asset
      2. Projeções GBM  — daily mean price per asset (downsampled)
      3. Métricas       — portfolio KPIs
    """
    wb = Workbook()

    # ── Sheet 1: Portfolio ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Portfólio RL"
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 14
    ws1.column_dimensions["H"].width = 18

    ws1.cell(row=1, column=1, value=f"Dashboard B3 – 18 Ativos Ibovespa – Perfil: {profile.capitalize()}")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws1.cell(row=2, column=1, value=f"Capital inicial: R$ {initial_capital:,.0f}  |  Algoritmo: Cross-Entropy Method (RL)")
    ws1.cell(row=2, column=1).font = Font(italic=True, color="888888")

    _header_row(ws1, 4, ["Ticker", "Setor", "Peso RL (%)", "Alocação (R$)",
                          "μ a.a. (%)", "σ a.a. (%)", "Ret. Esp. (%)", "Sharpe implícito"])
    assets_map = {a["ticker"]: a for a in sim["assets"]}
    for r_idx, (ticker, weight) in enumerate(zip(rl_result.tickers, rl_result.weights), start=5):
        asset = assets_map.get(ticker, {})
        alloc = initial_capital * float(weight)
        mu    = asset.get("mu", 0) * 100
        sigma = asset.get("sigma", 0) * 100
        ret_  = asset.get("terminal", {}).get("expected_return_pct", 0)
        sharpe_i = mu / max(sigma, 0.01)

        row_vals = [ticker, asset.get("sector", ""), round(weight * 100, 2),
                    round(alloc, 2), round(mu, 1), round(sigma, 1),
                    round(ret_, 2), round(sharpe_i, 3)]
        for c, val in enumerate(row_vals, start=1):
            ws1.cell(row=r_idx, column=c, value=val)
        # colour the return cell
        ret_cell = ws1.cell(row=r_idx, column=7)
        ret_cell.font = _GREEN_FG if ret_ >= 0 else _RED_FG

    # Summary box
    sum_row = 5 + len(rl_result.tickers) + 1
    ws1.cell(row=sum_row, column=1, value="Portfólio Total").font = Font(bold=True)
    ws1.cell(row=sum_row, column=3, value=100.0)
    ws1.cell(row=sum_row, column=4, value=round(initial_capital, 2))
    ws1.cell(row=sum_row, column=5, value=round(rl_result.portfolio_return * 100, 2))
    ws1.cell(row=sum_row, column=6, value=round(rl_result.portfolio_vol * 100, 2))
    ws1.cell(row=sum_row, column=8, value=round(rl_result.sharpe_ratio, 3))
    for c in range(1, 9):
        ws1.cell(row=sum_row, column=c).fill = _GOLD_FILL
        ws1.cell(row=sum_row, column=c).font = Font(bold=True, color="F59E0B")

    # ── Sheet 2: GBM Projections (downsampled to ~50 points) ─────────────────
    ws2 = wb.create_sheet("Projeções GBM")
    time_days = sim["time_days"]
    step = max(1, len(time_days) // 50)
    sampled_idx = list(range(0, len(time_days), step))
    if sampled_idx[-1] != len(time_days) - 1:
        sampled_idx.append(len(time_days) - 1)

    tickers = [a["ticker"] for a in sim["assets"]]
    _header_row(ws2, 1, ["Dia"] + tickers)
    ws2.column_dimensions["A"].width = 8
    for i, col_ticker in enumerate(tickers, start=2):
        ws2.column_dimensions[get_column_letter(i)].width = 10

    for r_idx, idx in enumerate(sampled_idx, start=2):
        ws2.cell(row=r_idx, column=1, value=time_days[idx])
        for c, asset in enumerate(sim["assets"], start=2):
            ws2.cell(row=r_idx, column=c, value=asset["mean"][idx])

    # ── Sheet 3: Risk Metrics ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Métricas de Risco")
    _header_row(ws3, 1, ["Métrica", "Valor"])
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 18
    metrics = [
        ("Perfil de Risco",          profile.capitalize()),
        ("Capital Inicial (R$)",     f"{initial_capital:,.2f}"),
        ("Retorno Esperado a.a.",     f"{rl_result.portfolio_return * 100:.2f}%"),
        ("Volatilidade a.a.",         f"{rl_result.portfolio_vol * 100:.2f}%"),
        ("Sharpe Ratio (RL)",         f"{rl_result.sharpe_ratio:.3f}"),
        ("Iterações CEM",             rl_result.iterations),
        ("N° de Ativos",              len(rl_result.tickers)),
        ("Algoritmo de Otimização",   "Cross-Entropy Method (CEM/RL)"),
        ("Simulações Monte Carlo",    sim["n_paths"]),
        ("Horizonte (anos)",          sim["T"]),
        ("",                          ""),
        ("Nota",                      "Modelo educacional – não é recomendação de investimento."),
    ]
    for r_idx, (k, v) in enumerate(metrics, start=2):
        ws3.cell(row=r_idx, column=1, value=k).font = Font(bold=bool(k))
        ws3.cell(row=r_idx, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
